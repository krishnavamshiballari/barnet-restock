import os
import re
import time
from datetime import date, timedelta
from pathlib import Path
import subprocess
import shutil
import pythoncom

import pandas as pd
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright
import win32com.client as win32  # Outlook COM
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ========= CONFIG =========
SENDER_SMTP = "krishna.ballari@astorialrs.ca"
BASE_URL   = "https://astoria.barnetportal.com"
REPORT_URL = f"{BASE_URL}/reports/sales-by-product"

# Write outputs into a repo-relative folder (works on Windows & Linux)
DOWNLOAD_DIR = Path(os.environ.get("DOWNLOAD_DIR", "out")).resolve()
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

ERROR_DIR = DOWNLOAD_DIR / "errors"
ERROR_DIR.mkdir(parents=True, exist_ok=True)


# Normal daily recipients (team)
EMAIL_TO = ["astorialiquorgeneral@gmail.com"]
EMAIL_CC = []

# ALERT recipient (you)
EMAIL_ALERT_TO = ["krishna.ballari@astorialrs.ca"]

START_TIME_STR  = "07:30 AM"
FINISH_TIME_STR = "11:50 PM"

KEEP_COLS   = ["SKU", "Description", "Category", "UOM", "Sold QTY"]
ENV_FILE    = ".env"
PROFILE_DIR = Path(".") / "barnet_profile"
PROFILE_DIR.mkdir(parents=True, exist_ok=True)


# For unattended runs via Task Scheduler:
HEADLESS = True
SLOW_MO  = 0
# ==========================


def log(msg: str):
    ts = time.strftime('%H:%M:%S')
    try:
        print(f"[{ts}] {msg}", flush=True)
    except UnicodeEncodeError:
        safe = ''.join(ch if ord(ch) < 128 else '?' for ch in msg)
        print(f"[{ts}] {safe}", flush=True)


def yesterday() -> date:
    return date.today() - timedelta(days=1)


def subject_for(_d: date) -> str:
    # subject uses TODAY's date (your preference), even though data is yesterday’s
    return f"{time.strftime('%m/%d/%Y')} Restocking List"


def output_clean_path(_d: date) -> Path:
    # filename uses TODAY's date to match your subject/body
    return DOWNLOAD_DIR / f"Restocking_{time.strftime('%Y-%m-%d')}.xlsx"


def email_body(_d: date) -> str:
    # body uses TODAY's date (your preference)
    return f"""Hi Team,

Please find attached restocking list for {time.strftime('%m/%d/%Y')}, let me know if you have any questions.

Thanks,
Krishna
"""


# ---------- Outlook helpers (COM) ----------
def _get_outlook_app():
    """
    Get a usable Outlook COM application even under Task Scheduler.
    """
    pythoncom.CoInitialize()  # required in scheduler/background
    # 1) Attach to running Outlook
    try:
        return win32.GetActiveObject("Outlook.Application")
    except Exception:
        pass
    # 2) Try to launch and then attach
    try:
        outlook_exe = shutil.which("outlook.exe")
        if outlook_exe:
            subprocess.Popen([outlook_exe, "/recycle"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            time.sleep(8)
            try:
                return win32.GetActiveObject("Outlook.Application")
            except Exception:
                pass
    except Exception as e:
        log(f"Note: couldn’t pre-launch Outlook: {repr(e)}")
    # 3) Create new COM server
    return win32.DispatchEx("Outlook.Application")

def _choose_account(app, smtp_address: str):
    """Return the Outlook account object matching the SMTP address, or None."""
    try:
        for acct in app.Session.Accounts:
            if str(getattr(acct, "SmtpAddress", "")).lower() == smtp_address.lower():
                return acct
    except Exception as e:
        log(f"Could not enumerate Outlook accounts: {repr(e)}")
    return None


# ---------- Playwright helpers ----------
def _first_visible(*locators):
    for loc in locators:
        try:
            if loc and loc.count() and loc.first.is_visible():
                return loc.first
        except Exception:
            continue
    return None


def click_apply(page) -> bool:
    log("Clicking Apply…")
    candidates = [
        page.get_by_role("button", name=re.compile(r"^\s*Apply\s*$", re.I)),
        page.locator("button:has-text('Apply')"),
        page.locator("input[type='button'][value='Apply']"),
        page.locator("input[type='submit'][value='Apply']"),
        page.locator("//button[normalize-space(.)='Apply']"),
        page.locator("text=Apply"),
    ]
    for sel in candidates:
        try:
            if sel and sel.count():
                btn = sel.first
                btn.scroll_into_view_if_needed()
                btn.click()
                page.wait_for_load_state("networkidle")
                time.sleep(1.5)
                log("Applied.")
                return True
        except Exception:
            continue
    log("Couldn’t find an Apply button.")
    return False


def ensure_logged_in(page, username: str | None, password: str | None):
    log("Opening Barnet portal…")
    page.goto(BASE_URL, wait_until="networkidle")

    user_loc = _first_visible(
        page.get_by_label(re.compile(r"^Username", re.I)),
        page.locator("xpath=//label[contains(.,'Username')]/following::input[1]"),
        page.locator("form input[type='text']").first,
        page.locator("input[name*=user i]").first,
    )
    pass_loc = _first_visible(
        page.get_by_label(re.compile(r"Access Code|Password", re.I)),
        page.locator("xpath=//label[contains(.,'Access Code') or contains(.,'Password')]/following::input[1]"),
        page.locator("form input[type='password']").first,
        page.locator("form input[type='text']").nth(1),
        page.locator("input[name*=pass i]").first,
    )
    login_btn = _first_visible(
        page.get_by_role("button", name=re.compile("login|sign in", re.I)),
        page.locator("xpath=//input[@type='submit' or @value='Login' or @value='LOG IN']"),
        page.locator("text=Login"),
    )

    if user_loc and pass_loc and login_btn:
        if not username or not password:
            raise RuntimeError("Creds missing. Put BARNET_USER/BARNET_PASS (or BARNET_USERNAME/BARNET_PASSWORD) in .env")
        log("Logging in…")
        for loc, val in [(user_loc, username), (pass_loc, password)]:
            loc.click()
            loc.press("Control+A")
            loc.type(str(val), delay=30)
        login_btn.click()
        page.wait_for_load_state("networkidle")
        time.sleep(1.0)
        log("Login completed.")
    else:
        log("No login form detected (likely already logged in).")


def select_store(page):
    try:
        page.wait_for_timeout(800)
        modal = page.locator("text=Select Store").first
        if modal and modal.count() and modal.is_visible():
            log("Selecting store: Astoria Liquor…")
            target = _first_visible(
                page.locator("text=Astoria Liquor 191222"),
                page.locator("text=Astoria Liquor"),
            )
            if target:
                target.click()
            try:
                _first_visible(
                    page.get_by_role("button", name=re.compile("Close", re.I)),
                    page.locator("button:has-text('Close')")
                )
            except Exception:
                pass
            page.wait_for_load_state("networkidle")
            time.sleep(0.8)
            log("Store selected.")
        else:
            log("Store selection modal not shown.")
    except Exception:
        log("Store selection step skipped (no modal).")


def set_report_filters(page, rep_date: date):
    log("Opening ‘Sales by Product’ report page…")
    page.goto(REPORT_URL, wait_until="networkidle")
    select_store(page)

    ds = rep_date.strftime("%m/%d/%Y")

    def clear_and_type(locator, text, press_enter=True):
        try:
            locator.click()
            locator.press("Control+A")
            locator.type(text, delay=30)
            if press_enter:
                locator.press("Enter")
            return True
        except Exception as e:
            log(f"  ! could not type into field: {e}")
            return False

    start_date = _first_visible(
        page.get_by_label(re.compile(r"^Start($|[^a-z])", re.I)),
        page.get_by_placeholder(re.compile(r"start|from", re.I)),
        page.locator("input[aria-label*='start' i], input[name*='start' i], input[id*='start' i]").first,
    )
    end_date = _first_visible(
        page.get_by_label(re.compile(r"^End($|[^a-z])|Finish Date", re.I)),
        page.get_by_placeholder(re.compile(r"end|finish|to", re.I)),
        page.locator("input[aria-label*='end' i], input[name*='end' i], input[id*='end' i]").first,
    )
    start_time = _first_visible(
        page.get_by_label(re.compile(r"Start Time|From Time", re.I)),
        page.get_by_placeholder(re.compile(r"time.*start|from time", re.I)),
        page.locator("input[aria-label*='start time' i], input[name*='starttime' i], input[id*='starttime' i]").first,
    )
    finish_time = _first_visible(
        page.get_by_label(re.compile(r"Finish Time|End Time|To Time", re.I)),
        page.get_by_placeholder(re.compile(r"time.*(finish|end|to)", re.I)),
        page.locator("input[aria-label*='finish time' i], input[name*='endtime' i], input[id*='endtime' i]").first,
    )

    log(f"  start_date field found: {bool(start_date)}")
    log(f"  end_date   field found: {bool(end_date)}")
    log(f"  start_time field found: {bool(start_time)}")
    log(f"  finish_time field found:{bool(finish_time)}")

    log(f"Setting date/time to {ds} {START_TIME_STR} – {ds} {FINISH_TIME_STR} …")
    if start_date:  clear_and_type(start_date, ds)
    if end_date:    clear_and_type(end_date, ds)
    if start_time:  clear_and_type(start_time, START_TIME_STR)
    if finish_time: clear_and_type(finish_time, FINISH_TIME_STR)

    clicked = click_apply(page)
    if not clicked:
        try:
            page.keyboard.press("Enter")
            page.wait_for_load_state("networkidle")
            time.sleep(1.0)
            log("Applied via Enter key.")
        except Exception:
            pass
    log("Filters applied.")


def export_to_excel(page) -> Path:
    log("Exporting to Excel (can take ~30–120 seconds)…")
    with page.expect_download(timeout=120_000) as d_info:
        try:
            page.get_by_role("button", name=re.compile("Export to Excel", re.I)).click()
        except:
            page.locator("text=Export to Excel").first.click()
    download = d_info.value
    suggested = download.suggested_filename or f"SalesByProduct_{yesterday().isoformat()}.xlsx"
    dest = DOWNLOAD_DIR / suggested
    download.save_as(str(dest))
    log(f"Downloaded: {dest}")
    return dest


# ---------- clean (grouped layout with totals & styling) ----------
def clean_and_reduce(in_path: str, out_path: str):
    """
    Build a grouped report that matches your manual format and styling.
    Returns (rows_total, rows_with_qty_gt0)
    """
    log("Cleaning and rebuilding grouped report…")

    # 1) Find header row
    raw = pd.read_excel(in_path, header=None, engine="openpyxl")
    raw = raw.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))

    REQUIRED = ["SKU", "Description", "Category", "Group", "UOM", "Sold QTY"]
    header_row_idx = None
    for i in range(min(40, len(raw))):
        row_vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        if all(any(h.lower() == v.lower() for v in row_vals) for h in REQUIRED):
            header_row_idx = i
            break
    if header_row_idx is None:
        header_row_idx = 0

    df = pd.read_excel(in_path, header=header_row_idx, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # 2) Normalize names
    alias = {"Sold Qty": "Sold QTY", "Qty Sold": "Sold QTY", "Unit": "UOM"}
    for k, v in alias.items():
        if k in df.columns and v not in df.columns:
            df.rename(columns={k: v}, inplace=True)

    missing = [c for c in ["SKU", "Description", "Category", "UOM", "Sold QTY"] if c not in df.columns]
    if missing:
        log(f"Missing expected columns: {missing}. Found: {df.columns.tolist()}")
        pd.DataFrame(columns=["SKU", "Description", "Category", "UOM", "Sold QTY"]).to_excel(out_path, index=False)
        return 0, 0

    # 3) Drop non-product lines
    sku_str = df["SKU"].astype(str).str.strip()
    is_blank = sku_str.eq("") | sku_str.eq("nan")
    is_total = sku_str.str.startswith("Total", na=False)
    is_section_header = sku_str.str.contains(r"^[A-Z].* - .*$", na=False) & df["Description"].isna()
    prod = df[~(is_blank | is_total | is_section_header)].copy()

    prod = prod[["SKU", "Description", "Category", "UOM", "Sold QTY"]].copy()

    def to_num(x):
        if isinstance(x, str):
            x = x.replace("$", "").replace(",", "").strip()
        return pd.to_numeric(x, errors="coerce")

    prod["Sold QTY"] = prod["Sold QTY"].apply(to_num).fillna(0).astype(int)

    def clean_sku(x):
        try:
            if isinstance(x, float) and x.is_integer():
                return int(x)
        except Exception:
            pass
        return str(x).strip()

    prod["SKU"] = prod["SKU"].apply(clean_sku)

    # guard null/blank categories
    prod = prod[prod["Category"].notna()].copy()
    prod["Category"] = prod["Category"].astype(str).str.strip()
    prod = prod[prod["Category"] != ""].copy()

    rows_total = len(prod)
    rows_gt0   = int((prod["Sold QTY"] > 0).sum())

    # 4) Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales By Product"

    # Title block
    title_font = Font(bold=True, size=12)
    small_font = Font(size=10)
    ws.append(["Sales By Product"])
    ws.append(["Astoria Liquor"])
    ws.append([yesterday().strftime("%m/%d/%y")])  # this line mirrors your manual "report date" line
    ws.append([f"Created: {time.strftime('%m/%d/%Y %I:%M %p')}"])
    ws["A1"].font = title_font
    ws["A2"].font = title_font
    ws["A3"].font = small_font
    ws["A4"].font = small_font
    ws.append([""])

    # Styles
    header_font = Font(bold=True)
    cat_font = Font(bold=True)
    total_font = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    start_row = ws.max_row + 1
    headers = ["SKU", "Description", "Category", "UOM", "Sold QTY"]

    # Keep first-seen category order
    codes, _uniques = pd.factorize(prod["Category"], sort=False, use_na_sentinel=False)
    prod["_cat_order"] = codes
    prod.sort_values(by=["_cat_order"], kind="stable", inplace=True)

    for category, grp in prod.groupby("Category", sort=False):
        # Category header
        ws.append([category])
        ws[ws.max_row][0].font = cat_font

        # Column headers
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        # Rows
        for _, r in grp.iterrows():
            ws.append([r["SKU"], r["Description"], r["Category"], r["UOM"], int(r["Sold QTY"])])
            row_i = ws.max_row
            ws[f"A{row_i}"].alignment = right
            ws[f"E{row_i}"].alignment = right
            for col in range(1, 6):
                ws[f"{get_column_letter(col)}{row_i}"].border = border

        # Total row
        total_qty = int(grp["Sold QTY"].sum())
        ws.append([f"Total {category}", "", "", "", total_qty])
        row_i = ws.max_row
        ws[f"A{row_i}"].font = total_font
        ws[f"E{row_i}"].font = total_font
        ws[f"E{row_i}"].alignment = right
        for col in range(1, 6):
            ws[f"{get_column_letter(col)}{row_i}"].border = border

        ws.append([""])  # spacer

    # Freeze panes at the first category section
    ws.freeze_panes = None

    # Column widths (auto-ish)
    base = {1: 8, 2: 28, 3: 12, 4: 8, 5: 10}
    for col_idx in range(1, 6):
        max_len = base.get(col_idx, 10)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)

    # Adjust column A width
    ws.column_dimensions['A'].width = 13.57

    wb.save(out_path)
    log(f"Saved formatted grouped report: {out_path}")
    return rows_total, rows_gt0


# ---------- email (COM via Outlook) ----------
def send_email(out_path: str, report_date: date):
    """
    Team email via Outlook COM, forced to send from SENDER_SMTP.
    """
    try:
        log("Connecting to Outlook (COM)…")
        app = _get_outlook_app()
        mail = app.CreateItem(0)  # olMailItem

        # Force the sender account
        acct = _choose_account(app, SENDER_SMTP)
        if acct:
            try:
                mail.SendUsingAccount = acct
            except Exception:
                # Fallback for some Outlook/pywin32 builds
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, acct))
        else:
            log(f"Couldn’t find Outlook account {SENDER_SMTP}. Using Outlook’s default sending account.")

        mail.To = ";".join(EMAIL_TO)
        if EMAIL_CC:
            mail.CC = ";".join(EMAIL_CC)
        mail.Subject = subject_for(report_date)
        mail.Body = email_body(report_date)
        if out_path and os.path.exists(out_path):
            mail.Attachments.Add(out_path)
        mail.Send()
        log("Email sent (Outlook).")
    except Exception as e:
        log(f"Outlook send failed: {repr(e)}")
        # Save draft so you can send manually
        try:
            mail = app.CreateItem(0)
            # keep same sender if possible
            acct = _choose_account(app, SENDER_SMTP)
            if acct:
                try:
                    mail.SendUsingAccount = acct
                except Exception:
                    mail._oleobj_.Invoke(*(64209, 0, 8, 0, acct))
            mail.To = ";".join(EMAIL_TO)
            if EMAIL_CC:
                mail.CC = ";".join(EMAIL_CC)
            mail.Subject = "[DRAFT] " + subject_for(report_date)
            mail.Body = email_body(report_date) + "\n\n(Saved as draft due to COM send error.)"
            if out_path and os.path.exists(out_path):
                mail.Attachments.Add(out_path)
            mail.Save()
            log("Saved message in Outlook Drafts.")
        except Exception as e2:
            log(f"Couldn’t save to Drafts: {repr(e2)}")


def send_alert_email(reason: str, out_path: str, report_date: date):
    """
    Alert email to you via Outlook COM, forced to send from SENDER_SMTP.
    """
    try:
        log("Connecting to Outlook (COM) for alert…")
        app = _get_outlook_app()
        mail = app.CreateItem(0)

        # Force the sender account
        acct = _choose_account(app, SENDER_SMTP)
        if acct:
            try:
                mail.SendUsingAccount = acct
            except Exception:
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, acct))
        else:
            log(f"Couldn’t find Outlook account {SENDER_SMTP}. Using Outlook’s default sending account.")

        mail.To = ";".join(EMAIL_ALERT_TO)
        mail.Subject = f"ALERT: Empty Restocking Report for {report_date.strftime('%m/%d/%Y')}"
        mail.Body = (
            f"Hi Krishna,\n\n"
            f"The restocking report for {report_date.strftime('%m/%d/%Y')} appears to be EMPTY.\n"
            f"Reason: {reason}\n\n"
            f"I've attached the cleaned file for reference.\n\n"
            f"— Auto Barnet Script"
        )
        if out_path and os.path.exists(out_path):
            mail.Attachments.Add(out_path)
        mail.Send()
        log("Alert email sent (Outlook).")
    except Exception as e:
        log(f"Outlook alert send failed: {repr(e)}")
        try:
            mail = app.CreateItem(0)
            acct = _choose_account(app, SENDER_SMTP)
            if acct:
                try:
                    mail.SendUsingAccount = acct
                except Exception:
                    mail._oleobj_.Invoke(*(64209, 0, 8, 0, acct))
            mail.To = ";".join(EMAIL_ALERT_TO)
            mail.Subject = "[DRAFT] " + f"ALERT: Empty Restocking Report for {report_date.strftime('%m/%d/%Y')}"
            mail.Body = (
                f"(Saved as draft due to COM send error)\n\n"
                f"Reason: {reason}\n"
            )
            if out_path and os.path.exists(out_path):
                mail.Attachments.Add(out_path)
            mail.Save()
            log("Saved alert in Outlook Drafts.")
        except Exception as e2:
            log(f"Couldn’t save alert to Drafts: {repr(e2)}")

# --- CI guard: skip Outlook COM email when running in GitHub Actions ---
IS_CI = os.getenv("CI", "").lower() == "true" or os.getenv("DRY_RUN_EMAIL") == "1"

# ---------- main ----------
def main():
    # Read .env (supports either USER/PASS or USERNAME/PASSWORD)
    DOTENV_PATH = Path(__file__).parent / ENV_FILE
    load_dotenv(DOTENV_PATH)
    username = os.getenv("BARNET_USER") or os.getenv("BARNET_USERNAME")
    password = os.getenv("BARNET_PASS") or os.getenv("BARNET_PASSWORD")

    rep_date  = yesterday()  # always pull yesterday’s data
    out_clean = output_clean_path(rep_date)

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=HEADLESS,
            slow_mo=(SLOW_MO if not HEADLESS else 0),
            accept_downloads=True,
            base_url=BASE_URL,
            viewport={"width": 1400, "height": 900},
        )
        page = context.new_page()

        ensure_logged_in(page, username, password)
        select_store(page)
        set_report_filters(page, rep_date)

        # extra safety: Apply again before exporting
        click_apply(page)

        downloaded = export_to_excel(page)
        rows_total, rows_gt0 = clean_and_reduce(str(downloaded), str(out_clean))
        context.close()

    # Decide: normal send vs alert
    if rows_total == 0:
        if IS_CI:
            log("[CI] Skipping alert email (rows_total == 0).")
        else:
            send_alert_email("No product rows found after cleaning.", str(out_clean), rep_date)
    elif rows_gt0 == 0:
        if IS_CI:
            log("[CI] Skipping alert email (all Sold QTY == 0).")
        else:
            send_alert_email("All products have Sold QTY = 0.", str(out_clean), rep_date)
    else:
        if IS_CI:
            log("[CI] Skipping normal email send.")
        else:
            send_email(str(out_clean), rep_date)

    log(f"Finished run for {rep_date}. Rows: {rows_total}, >0 qty: {rows_gt0}")


if __name__ == "__main__":
    main()
